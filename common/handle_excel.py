"""
Time:2022/8/2 000214:47
Author:bty
"""
import openpyxl
import os
from openpyxl.styles import PatternFill

from common.handle_path import DATAS_PATH


class HandleExcel:
    def __init__(self, file_name: str, sheet_name: str):
        """

        :param file_name: excel文件目录
        :param sheet_name: excel文件中的sheet名称
        """
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_excel(self) -> list:
        """
        读取excel文件，将案例生成列表的形式返回
        :return:
        """
        workbook = openpyxl.load_workbook(filename=self.file_name)
        sheet = workbook[self.sheet_name]
        res = list(sheet.values)
        key = [key for key in res[0]]
        case = []
        for values in res[1:]:
            v = [v for v in values]
            case.append(dict(zip(key, v)))
        return case

    def write_excel(self, row, column, value, colour=True):
        """
        将返回的结果写入excel中
        :param row: 行
        :param column: 列
        :param value: 值
        :param colour: 是否需要填充颜，默认不需要填充颜色
        :return:
        """
        workbook = openpyxl.load_workbook(filename=self.file_name)
        sheet = workbook[self.sheet_name]
        fille = PatternFill('solid', fgColor='FF0000')  # 设置填充颜色为 红色：FF0000 橙色：FFBB00
        no_fille = PatternFill()  # 设置无填充色
        if colour:
            sheet.cell(row=row, column=column, value=value).fill = no_fille
        else:
            sheet.cell(row=row, column=column, value=value).fill = fille
        workbook.save(self.file_name)


if __name__ == '__main__':
    excel = HandleExcel(file_name=os.path.join(DATAS_PATH, 'newApiCase.xlsx'), sheet_name='register')
    print(excel.read_excel())
